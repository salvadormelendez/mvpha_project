#!/usr/bin/env python3

#####################################
# APPOINTMENT LETTER GENERATOR v1.0
# Designed and Written by S.M.
#####################################

import os
from datetime import date, datetime, timedelta
import calendar
import math
import docx2pdf
from PyPDF2 import PdfFileMerger
import time
from mailmerge import MailMerge
from openpyxl import Workbook
from openpyxl import load_workbook
from win32com.client import Dispatch


#GET DATE - mm_dd_yy
today = date.today()
d = today.strftime("%m_%d_%Y")
dt = datetime.now()
e = dt.strftime("%m_%d_%Y_%H_%M_%S")

#APPOINTMENTS WORKBOOK
excel_file = "appointments.xlsx"
#GET FILES IN DIRECTORY
current_dir = os.getcwd()
path_input = current_dir + "\\input\\"
path_date = current_dir + "\\output\\" + e + "\\"
path_word = current_dir + "\\output\\" + e + "\\word\\"
path_pdf = current_dir + "\\output\\" + e + "\\pdf\\"
#DATE FOLDERS
if not os.path.exists(path_date):
    os.mkdir(path_date)
    os.mkdir(path_word)
    os.mkdir(path_pdf)
#INPUT FILES
for files in os.listdir(path_input):
    if os.path.isfile(os.path.join(path_input, files)):
        input_file = files

#COPY EXCEL TEMPLATE (TABLE)
msg = 'copy ' + '"' + path_input + input_file + '" "' + path_date + excel_file + '"'
os.system(msg)
path_template = current_dir + "\\template.xlsx"
path_excel = path_date + excel_file
xl = Dispatch("Excel.Application")
wb1 = xl.Workbooks.Open(Filename=path_template)
wb2 = xl.Workbooks.Open(Filename=path_excel)
ws1 = wb1.Worksheets(1)
ws1.Copy(Before=wb2.Worksheets(1))
wb2.Close(SaveChanges=True)
xl.Quit()

#MODIFY EXCEL WORKBOOK
wb = load_workbook(path_excel)
ws1 = wb['wlAllTenants']
ws2 = wb['1st_appt']
end_row = 8
for row in range(8,ws1.max_row+1):
    if(ws1.cell(row,1).value is None or "Total Applicants:" in ws1.cell(row,1).value):
        break
    end_row = row
#EXTRACT DATA FROM WORKSHEET
name = []
address = []
city = []
phone = []
date = []
beds = []
for row in range(8,end_row+1):
	mod_name = ws1.cell(row,1).value.title()
	name.append(mod_name)
	mod_address = ws1.cell(row,6).value.title()
	address.append(mod_address)
	city.append(ws1.cell(row,9).value)
	phone.append(ws1.cell(row,10).value)
	date.append(ws1.cell(row,11).value)
	beds.append(ws1.cell(row,14).value)
#FILL SUMMARY TABLE
for i in range(len(name)):
	ws2.cell(i+2,1).value = name[i]
	ws2.cell(i+2,2).value = address[i]
	ws2.cell(i+2,3).value = city[i]
	ws2.cell(i+2,4).value = date[i]
	ws2.cell(i+2,5).value = beds[i]
#ADJUST TABLE SIZE
msg = "A1:H" + str(len(name)+1)
ws2.tables['Table1'].ref = msg
#GENERATE APPOINTMENTS - DATE & TIME
weekdays = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
todays_name = calendar.day_name[today.weekday()]
todays_index = weekdays.index(todays_name)
appt_times = ['8:00 AM', '8:30 AM', '9:00 AM', '9:30 AM', '10:00 AM', '10:30 AM', '11:00 AM', '11:30 AM', '1:30 PM', '2:00 PM', '2:30 PM', '3:00 PM', '3:30 PM']
td = timedelta(14)
#SET STARTING APPOINTMENT DATE
start_appt_date = today + td
start_appt_date_name = calendar.day_name[start_appt_date.weekday()]
while(True):
	if start_appt_date_name not in weekdays:
		start_appt_date = start_appt_date+timedelta(1)
		start_appt_date_name = calendar.day_name[start_appt_date.weekday()]
	else:
		break
#CALCULATE NUMBER OF APPOINTMENT DATES
total_days = math.ceil(len(name)/len(appt_times))
appt_dates = []
appt_formal_en = []
appt_days = []
for i in range(0,total_days):
	possible_appt_date = start_appt_date+timedelta(i)
	possible_name = calendar.day_name[possible_appt_date.weekday()]
	while(True):
		if possible_name in weekdays:
			str_date = possible_appt_date.strftime("%m/%d/%Y")
			appt_dates.append(str_date)
			appt_days.append(possible_name)
			formal_date = possible_appt_date.strftime("%A %B %d, %Y")
			appt_formal_en.append(formal_date)
			break
		else:
			possible_appt_date = possible_appt_date+timedelta(1)
			possible_name = calendar.day_name[possible_appt_date.weekday()]
#GENERATE DATES IN SPANISH
appt_formal_es = []
dict_weekdays = {'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miercoles', 'Thursday': 'Jueves', 'Friday': 'Viernes'}
dict_months = {'January': 'Enero', 'February': 'Febrero', 'March': 'Marzo', 'April': 'Abril', 'May': 'Mayo', 'June': 'Junio', 'July': 'Julio', 'August': 'Agosto', 'September': 'Septiembre', 'October': 'Octubre', 'November': 'Noviembre', 'December': 'Diciembre'}
for i in appt_formal_en:
	for english, spanish in dict_weekdays.items():
		i = i.replace(english, spanish)
	for english, spanish in dict_months.items():
		i = i.replace(english, spanish)
	appt_formal_es.append(i)
#SET APPOINTMENT DATES & TIMES
appt_day = []
appt_date = []
appt_en = []
appt_es = []
appt_time = []
for i in range(len(name)):
	#SET DATE & DAY
	d_index = math.floor(i/len(appt_times))
	appt_date.append(appt_dates[d_index])
	appt_en.append(appt_formal_en[d_index])
	appt_es.append(appt_formal_es[d_index])
	appt_day.append(appt_days[d_index])
	#SET TIME
	appt_time.append(appt_times[i-(13*d_index)])
	#WRITE TO CELLS
	ws2.cell(i+2,6).value = appt_day[i]
	ws2.cell(i+2,7).value = appt_date[i]
	ws2.cell(i+2,8).value = appt_time[i]
#SAVE AND CLOSE
wb.save(path_excel)
wb.close()


#WORD
#CREATE FILE NAMES
word_files = []
for i in range(len(name)):
	aux = name[i].split(" ")
	w_file = aux[-1] + "_" + aux[0]
	word_files.append(w_file)
	with MailMerge('template.docx') as document:
		w_name = str(name[i])
		w_address = str(address[i])
		w_city = str(city[i])
		w_beds = str(beds[i])
		w_date = str(date[i])
		w_appt_en = str(appt_en[i])
		w_appt_es = str(appt_es[i])
		w_time = str(appt_time[i])
		document.merge(name=w_name, address=w_address, city=w_city, beds=w_beds, appdate=w_date, aptdateen=w_appt_en, aptdatees=w_appt_es, time=w_time)
		output_file = path_date + word_files[i] + "_PH_1st_screen_appt.docx"
		document.write(output_file)
		#CONVERT ALL WORD FILES TO PDF FILES
		docx2pdf.convert(output_file)
#MOVE ALL WORD FILES TO "word" FOLDER
original = path_date + "*.docx"
msg = 'move ' + '"' + original + '" "' + path_word + '"'
os.system(msg)
#MOVE ALL PDF FILES TO "pdf" FOLDER
original = path_date + "*.pdf"
msg = 'move ' + '"' + original + '" "' + path_pdf + '"'
os.system(msg)
#MERGE ALL PDF FILES
os.chdir(path_pdf)
x = [a for a in os.listdir() if a.endswith(".pdf")]
merger = PdfFileMerger()
for pdf in x:
    merger.append(open(pdf, 'rb'))
output_pdf = path_date + "all_appointment_letters.pdf"
with open(output_pdf, "wb") as fout:
    merger.write(fout)
